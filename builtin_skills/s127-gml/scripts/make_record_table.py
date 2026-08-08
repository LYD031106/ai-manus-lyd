#!/usr/bin/env python3
"""从要素清单 JSON 生成《S-127要素生产与检查记录表》。

指南 5.2 要求：GML 的 text 里不写条款序号，但记录表必须记录条款来源以便溯源；
指南 3.3.2(4)(5) 要求：GML 无法承载的关联类型、以及一审/二审意见都在记录表里流转。

用法：
    python make_record_table.py featureset.json -o 记录表.xlsx
    python make_record_table.py featureset.json -o 记录表      # 无 openpyxl 时输出 CSV

产出三个工作表（与生产线现行表式一致）：
    文件解析 / gml校对 / gml审核
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from s127_model import GEOGRAPHIC_FEATURES, REUSABLE_BY_DEFAULT, ROLE_TITLE  # noqa: E402

COLUMNS = [
    "地理要素类型",
    "地理要素名称",
    "地理要素出处",
    "信息要素类型",
    "信息要素名称",
    "信息要素出处",
    "关联要素及关联类型",
    "修改意见",
]
SHEETS = ["文件解析", "gml校对", "gml审核"]


def label(feat: dict) -> str:
    fn = feat.get("featureName") or {}
    if isinstance(fn, dict):
        return " ".join(x for x in (fn.get("zho"), fn.get("eng")) if x)
    if isinstance(fn, list):
        return " ".join(str(x.get("name", "")) for x in fn)
    return str(feat.get("ref", ""))


def source_of(feat: dict, common: dict) -> str:
    reusable = feat.get("reusable", feat.get("type") in REUSABLE_BY_DEFAULT)
    si = feat.get("sourceIndication") or ({} if reusable else common.get("sourceIndication") or {})
    if reusable and not si:
        return "/"  # 可复用要素不填来源（指南 4.3）
    src = si.get("source")
    if isinstance(src, (list, tuple)):
        src = " ".join(str(s) for s in src)
    parts = [str(src)] if src else []
    if feat.get("clause"):
        parts.append(str(feat["clause"]))
    return "；".join(parts) or "/"


def assoc_text(feat: dict, refs: dict) -> str:
    out = []
    for a in feat.get("associations") or []:
        role = a.get("role", "?")
        title = ROLE_TITLE.get(role, role)
        target = refs.get(a.get("target"), {})
        kind = a.get("permission") or a.get("inclusion")
        piece = f"{role}→{label(target) or a.get('target')}"
        if title in ("PermissionType", "InclusionType"):
            piece += f"（{kind or '待补录'}）"
        out.append(piece)
    return "\n".join(out)


def build_rows(doc: dict) -> list[list[str]]:
    common = doc.get("common") or {}
    features = doc.get("features") or []
    refs = {(f.get("ref") or f.get("id")): f for f in features}
    geo = [f for f in features if f.get("type") in GEOGRAPHIC_FEATURES]
    info = [f for f in features if f.get("type") not in GEOGRAPHIC_FEATURES]

    rows: list[list[str]] = []
    for i in range(max(len(geo), len(info))):
        g = geo[i] if i < len(geo) else None
        n = info[i] if i < len(info) else None
        rows.append(
            [
                g["type"] if g else "",
                label(g) if g else "",
                source_of(g, common) if g else "",
                n["type"] if n else "",
                label(n) if n else "",
                source_of(n, common) if n else "",
                assoc_text(g, refs) if g else "",
                "",
            ]
        )
    return rows


def write_xlsx(path: Path, rows: list[list[str]], meta: dict) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return False

    wb = Workbook()
    for idx, sheet in enumerate(SHEETS):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet
        ws.append(COLUMNS)
        head_fill = PatternFill("solid", fgColor="DDEBF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in rows:
            ws.append(row if sheet == "文件解析" else row[:-1] + [""])
        widths = [26, 46, 34, 22, 30, 22, 40, 34]
        for col, width in zip("ABCDEFGH", widths):
            ws.column_dimensions[col].width = width
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        note = f"数据集：{meta.get('name', '')}    要素总数：{len(rows)}    阶段：{sheet}"
        ws.cell(row=len(rows) + 3, column=1, value=note)
    wb.save(path)
    return True


def write_csv(base: Path, rows: list[list[str]]) -> list[Path]:
    written = []
    for sheet in SHEETS:
        path = base.with_name(f"{base.stem}-{sheet}.csv")
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(COLUMNS)
            writer.writerows(rows if sheet == "文件解析" else [r[:-1] + [""] for r in rows])
        written.append(path)
    return written


def main() -> int:
    ap = argparse.ArgumentParser(description="生成 S-127 要素生产与检查记录表")
    ap.add_argument("featureset")
    ap.add_argument("-o", "--output", default="S-127要素生产与检查记录表.xlsx")
    args = ap.parse_args()

    doc = json.loads(Path(args.featureset).read_text(encoding="utf-8"))
    rows = build_rows(doc)
    out = Path(args.output)

    if out.suffix.lower() == ".xlsx" and write_xlsx(out, rows, doc.get("dataset") or {}):
        print(f"[OK] 已生成 {out}（{len(rows)} 行 × 3 个工作表）")
    else:
        if out.suffix.lower() == ".xlsx":
            print("[提醒] 未安装 openpyxl，改为输出 CSV（pip install openpyxl 可直接出 xlsx）")
        for path in write_csv(out, rows):
            print(f"[OK] 已生成 {path}")
    print("[下一步] 把条款出处（第X条/附件X）填进「地理要素出处」列，并按一审/二审意见填「修改意见」")
    return 0


if __name__ == "__main__":
    sys.exit(main())
